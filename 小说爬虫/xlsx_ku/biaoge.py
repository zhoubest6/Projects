import openpyxl

def xie(path,data):
    wb = openpyxl.load_workbook(path)
    li = wb.sheetnames
    sheet = wb.get_sheet_by_name(li[0])  # 获取工作表
    s=0
    s+=len(data[1:])
    for i in range(1,len(data)):
        sheet.append([i, data[i-1]])# 写入xlsx表
        # if s==i:
        #     sheet.append(["下一章：", "___________________内容________________"])  # 写入xlsx表
    wb.save(path)
    wb.close()
def du(path):
    wb = openpyxl.load_workbook(path)
    li = wb.sheetnames
    sheet = wb.get_sheet_by_name(li[0])  # 获取工作表
    max_h = sheet.max_row  # 获取最大行数
    zhi=[]
    for i in range(1, max_h):
        #print(sheet.cell(i, 1).value,sheet.cell(i, 2).value)
        s=sheet.cell(i, 2).value
        zhi.append(s)
    return zhi
    # 获取单元格A1值，column与row依然可用
def shan_chu(path):
    wb = openpyxl.load_workbook(path)
    li = wb.sheetnames
    sheet = wb.get_sheet_by_name(li[0])  # 获取工作表
    max_h = sheet.max_row  # 获取最大行数
    for i in range(1, (max_h+1)):
        del sheet["B{}".format(i)]
        del sheet["A{}".format(i)]
    wb.save(path)
    print("__全删除__%s行"%max_h)
    wb.close()




